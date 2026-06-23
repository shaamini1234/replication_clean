import os
import logging
import re
from typing import Any, Dict, List, Optional

import openpyxl
import pandas as pd

from .schema import (
    BalanceOfPayments,
    DebtInterest,
    DetailedReceipts,
    FiscalAggregates,
    GDPComponents,
    GDPIncomeComponents,
    HouseholdBalanceSheet,
    HouseholdIncome,
    LabourMarket,
    MarketAssumptions,
    NominalGDPComponents,
    OutputGap,
    PolicyMeasure,
    PotentialOutput,
    PriceIndices,
    ReceiptsBreakdown,
)

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# OBR Forecast Loader (vintage-based)
# ---------------------------------------------------------------------------

class OBRForecastLoader:
    """
    Loads OBR Detailed Forecast Tables from a vintage directory.

    Handles both naming conventions:
      - March 2025:  Aggregates_Detailed_forecast_tables_March_2025.xlsx
      - March 2026:  efo-march-2026-detailed-forecast-tables-aggregates.xlsx
    """

    def __init__(self, vintage_dir: str):
        self.vintage_dir = vintage_dir
        self.files = self._detect_files()

    def _detect_files(self) -> Dict[str, str]:
        files = {}
        if not os.path.isdir(self.vintage_dir):
            logger.warning("Vintage directory not found: %s", self.vintage_dir)
            return files

        for fname in os.listdir(self.vintage_dir):
            lower = fname.lower()
            if not (lower.endswith('.xlsx') or lower.endswith('.xls')):
                continue
            path = os.path.join(self.vintage_dir, fname)
            if 'aggregates' in lower and ('detailed' in lower or 'forecast' in lower):
                files['aggregates'] = path
            elif 'economy' in lower and ('detailed' in lower or 'forecast' in lower):
                files['economy'] = path
            elif 'receipts' in lower and ('detailed' in lower or 'forecast' in lower):
                files['receipts'] = path
            elif 'expenditure' in lower and ('detailed' in lower or 'forecast' in lower):
                files['expenditure'] = path
            elif 'debt' in lower and 'interest' in lower:
                files['debt_interest'] = path
            elif 'policy' in lower and ('detailed' in lower or 'forecast' in lower):
                files['policy'] = path
        return files

    # -- Generic quarterly parser ------------------------------------------

    def _parse_quarterly_sheet(
        self,
        file_key: str,
        sheet_name: str,
        header_map: Dict[str, str],
        header_rows: Optional[List[int]] = None,
        data_start_row: int = 4,
        col_range: Optional[range] = None,
    ) -> List[Dict[str, Any]]:
        """
        Parse a quarterly OBR Economy sheet into a list of dicts.

        Args:
            file_key: Key in self.files ('economy', 'aggregates', etc.)
            sheet_name: Excel sheet name (e.g. '1.1')
            header_map: Maps header text prefixes (lowercase) to field names.
                        Longer prefixes are matched first to avoid ambiguity.
            header_rows: Row indices to scan for column headers. Default [3].
            data_start_row: First row that may contain quarterly data.

        Returns:
            List of dicts with 'date' key plus mapped field values.
        """
        if header_rows is None:
            header_rows = [3]
        if col_range is None:
            col_range = range(3, 25)

        filepath = self.files.get(file_key)
        if not filepath:
            logger.warning("No %s file found in %s", file_key, self.vintage_dir)
            return []

        wb = openpyxl.load_workbook(filepath, data_only=True)
        if sheet_name not in wb.sheetnames:
            logger.warning("Sheet '%s' not found in %s", sheet_name, filepath)
            wb.close()
            return []
        ws = wb[sheet_name]

        # Map columns to field names from header rows
        sorted_patterns = sorted(
            header_map.items(), key=lambda x: len(x[0]), reverse=True,
        )
        col_map: Dict[int, str] = {}
        assigned: set = set()

        for hr in header_rows:
            for col in col_range:
                raw = ws.cell(row=hr, column=col).value
                if raw is None:
                    continue
                label = str(raw).strip().lower()
                for pattern, field in sorted_patterns:
                    if field not in assigned and label.startswith(pattern):
                        col_map[col] = field
                        assigned.add(field)
                        break

        if not col_map:
            logger.warning(
                "No headers matched in sheet '%s' of %s", sheet_name, filepath,
            )
            wb.close()
            return []

        # Read quarterly data rows
        records: List[Dict[str, Any]] = []
        empty_run = 0
        row = data_start_row
        while empty_run < 10:
            date_val = ws.cell(row=row, column=2).value
            date_str = str(date_val).strip() if date_val is not None else ''
            if re.match(r'\d{4}Q[1-4]', date_str):
                empty_run = 0
                record: Dict[str, Any] = {'date': date_str}
                for col, field in col_map.items():
                    val = ws.cell(row=row, column=col).value
                    try:
                        record[field] = float(val) if val is not None else 0.0
                    except (ValueError, TypeError):
                        record[field] = 0.0
                records.append(record)
            else:
                empty_run += 1
            row += 1

        wb.close()
        logger.info(
            "Loaded %d quarters from %s sheet '%s'",
            len(records), file_key, sheet_name,
        )
        return records

    # -- Fiscal aggregates (Table 6.5) ------------------------------------

    def load_fiscal_aggregates(self) -> List[FiscalAggregates]:
        """
        Parse Aggregates Table 6.5: Components of net borrowing.

        Layout (consistent across vintages):
            Row 5:  fiscal year headers in columns C–H
            Row 7:  Current receipts
            Row 8:  Current expenditure
            Row 9:  Depreciation
            Row 10: Current budget surplus/deficit
            Row 12: Gross investment
            Row 14: Net investment
            Row 15: Net borrowing
        """
        filepath = self.files.get('aggregates')
        if not filepath:
            logger.warning("No aggregates file found in %s", self.vintage_dir)
            return []

        logger.info("Loading fiscal aggregates from %s", filepath)
        wb = openpyxl.load_workbook(filepath, data_only=True)

        sheet_name = '6.5'
        if sheet_name not in wb.sheetnames:
            logger.warning("Sheet '%s' not found in %s", sheet_name, filepath)
            wb.close()
            return []
        ws = wb[sheet_name]

        # Row 5: fiscal year headers starting from column C
        fiscal_years = []
        col = 3  # column C
        while True:
            val = ws.cell(row=5, column=col).value
            if val is None:
                break
            fiscal_years.append(str(val).strip())
            col += 1

        if not fiscal_years:
            logger.warning("No fiscal years found in row 5")
            wb.close()
            return []

        # Read data rows by label (search rows 6-20 to be robust)
        label_rows = {}
        LABELS = {
            'current receipts': 'current_receipts',
            'current expenditure': 'current_expenditure',
            'depreciation': 'depreciation',
            'surplus on current budget': 'current_budget_surplus',
            'current budget surplus': 'current_budget_surplus',
            'current budget deficit': 'current_budget_deficit',
            'gross investment': 'gross_investment',
            'net investment': 'net_investment',
            'net borrowing': 'net_borrowing',
        }
        for row_idx in range(6, 25):
            raw = ws.cell(row=row_idx, column=2).value
            if raw is None:
                continue
            label = str(raw).strip().lower().rstrip('1234567890 ')
            for search_text, field_name in LABELS.items():
                if label == search_text or label.startswith(search_text):
                    values = []
                    for i, _ in enumerate(fiscal_years):
                        cell_val = ws.cell(row=row_idx, column=3 + i).value
                        values.append(float(cell_val) if cell_val is not None else 0.0)
                    label_rows[field_name] = values
                    break

        wb.close()

        # Build FiscalAggregates objects
        results = []
        for i, fy in enumerate(fiscal_years):
            # Handle the deficit/surplus sign convention:
            #   "Surplus on current budget" is positive when in surplus
            #   "Current budget deficit" is positive when in deficit (= -surplus)
            if 'current_budget_surplus' in label_rows:
                cbs = label_rows['current_budget_surplus'][i]
            elif 'current_budget_deficit' in label_rows:
                cbs = -label_rows['current_budget_deficit'][i]
            else:
                cbs = 0.0

            results.append(FiscalAggregates(
                fiscal_year=fy,
                current_receipts=label_rows.get('current_receipts', [0.0] * len(fiscal_years))[i],
                current_expenditure=label_rows.get('current_expenditure', [0.0] * len(fiscal_years))[i],
                depreciation=label_rows.get('depreciation', [0.0] * len(fiscal_years))[i],
                current_budget_surplus=cbs,
                gross_investment=label_rows.get('gross_investment', [0.0] * len(fiscal_years))[i],
                net_investment=label_rows.get('net_investment', [0.0] * len(fiscal_years))[i],
                net_borrowing=label_rows.get('net_borrowing', [0.0] * len(fiscal_years))[i],
            ))

        logger.info("Loaded %d fiscal years: %s", len(results),
                     [r.fiscal_year for r in results])
        return results

    # -- Economy tables (quarterly) ----------------------------------------

    def load_gdp_components(self) -> List[GDPComponents]:
        """Parse Economy Table 1.1: GDP expenditure components (real, quarterly)."""
        header_map = {
            'private consumption': 'private_consumption',
            'government consumption': 'government_consumption',
            'fixed investment': 'fixed_investment',
            'business investment': 'business_investment',
            'private dwellings': 'private_dwellings',
            'general government': 'general_government_investment',
            'change in inventories': 'change_in_inventories',
            'net acquisition': 'valuables',
            'total final expenditure': 'total_final_expenditure',
            'exports': 'exports',
            'imports': 'imports',
            'statistical discrepancy': 'statistical_discrepancy',
            'real gdp': 'gdp',
            'non-oil gva': 'non_oil_gva',
        }
        records = self._parse_quarterly_sheet(
            'economy', '1.1', header_map,
            header_rows=[3, 4], data_start_row=5,
        )
        return [
            GDPComponents(
                date=r['date'],
                private_consumption=r.get('private_consumption', 0.0),
                government_consumption=r.get('government_consumption', 0.0),
                fixed_investment=r.get('fixed_investment', 0.0),
                business_investment=r.get('business_investment', 0.0),
                private_dwellings=r.get('private_dwellings', 0.0),
                general_government_investment=r.get('general_government_investment', 0.0),
                change_in_inventories=r.get('change_in_inventories', 0.0),
                exports=r.get('exports', 0.0),
                imports=r.get('imports', 0.0),
                gdp=r.get('gdp', 0.0),
                valuables=r.get('valuables', 0.0),
                total_final_expenditure=r.get('total_final_expenditure', 0.0),
                statistical_discrepancy=r.get('statistical_discrepancy', 0.0),
                non_oil_gva=r.get('non_oil_gva', 0.0),
            )
            for r in records
        ]

    def load_gdp_income(self) -> List[GDPIncomeComponents]:
        """Parse Economy Table 1.3: GDP income components (nominal, quarterly)."""
        header_map = {
            'labour income': 'labour_income',
            'non-oil pnfc': 'non_oil_pnfc_profits',
            'other income': 'other_income',
            'gross value added at factor cost': 'gva_factor_cost',
            'gva at factor cost': 'gva_factor_cost',
            'taxes on products': 'taxes_on_products_less_subsidies',
            'statistical discrepancy': 'statistical_discrepancy',
            'gdp at market prices': 'gdp_market_prices',
        }
        records = self._parse_quarterly_sheet(
            'economy', '1.3', header_map, data_start_row=4,
        )
        return [
            GDPIncomeComponents(
                date=r['date'],
                labour_income=r.get('labour_income', 0.0),
                non_oil_pnfc_profits=r.get('non_oil_pnfc_profits', 0.0),
                other_income=r.get('other_income', 0.0),
                gva_factor_cost=r.get('gva_factor_cost', 0.0),
                taxes_on_products_less_subsidies=r.get('taxes_on_products_less_subsidies', 0.0),
                statistical_discrepancy=r.get('statistical_discrepancy', 0.0),
                gdp_market_prices=r.get('gdp_market_prices', 0.0),
            )
            for r in records
        ]

    def load_labour_market(self) -> List[LabourMarket]:
        """Parse Economy Table 1.6: Labour market (quarterly)."""
        header_map = {
            'employment rate': 'employment_rate',
            'employment (': 'employment_millions',
            'employees (': 'employees_millions',
            'ilo unemployment rate': 'unemployment_rate',
            'ilo unemployment (': 'unemployment_millions',
            'participation rate': 'participation_rate',
            'average hours': 'average_hours',
            'total hours': 'total_hours_millions',
            'compensation of employees': 'compensation_of_employees',
            'wages and salaries': 'wages_salaries_bn',
            'average weekly earnings index': 'awe_index',
        }
        records = self._parse_quarterly_sheet(
            'economy', '1.6', header_map, data_start_row=4,
        )
        return [
            LabourMarket(
                date=r['date'],
                employment_millions=r.get('employment_millions', 0.0),
                employment_rate=r.get('employment_rate', 0.0),
                unemployment_millions=r.get('unemployment_millions', 0.0),
                unemployment_rate=r.get('unemployment_rate', 0.0),
                participation_rate=r.get('participation_rate', 0.0),
                average_hours=r.get('average_hours', 0.0),
                total_hours_millions=r.get('total_hours_millions', 0.0),
                compensation_of_employees=r.get('compensation_of_employees', 0.0),
                employees_millions=r.get('employees_millions', 0.0),
                wages_salaries_bn=r.get('wages_salaries_bn', 0.0),
                awe_index=r.get('awe_index', 0.0),
            )
            for r in records
        ]

    def load_output_gap(self) -> List[OutputGap]:
        """Parse Economy Table 1.14: OBR output gap estimate (quarterly, from 1972Q1).

        No header row — data starts at row 3 with date in col 2, value in col 3.
        """
        filepath = self.files.get('economy')
        if not filepath:
            return []
        wb = openpyxl.load_workbook(filepath, data_only=True)
        if '1.14' not in wb.sheetnames:
            wb.close()
            return []
        ws = wb['1.14']
        records = []
        for row in ws.iter_rows(min_row=3, values_only=True):
            date_val = row[1] if len(row) > 1 else None
            val = row[2] if len(row) > 2 else None
            if date_val and re.match(r'\d{4}Q[1-4]', str(date_val).strip()):
                try:
                    records.append(OutputGap(
                        date=str(date_val).strip(),
                        output_gap_pct=float(val) if val is not None else 0.0,
                    ))
                except (ValueError, TypeError):
                    pass
        wb.close()
        return records

    def load_potential_output(self) -> List[PotentialOutput]:
        """Parse Economy Table 1.15: Potential output forecast (quarterly, from 2019Q1).

        Col 4: Potential output (£m) → convert to £bn for TRGDP.
        Col 9: Equilibrium unemployment rate (NAIRU).
        Col 7: Potential employment rate.
        Col 10: Potential average hours.
        Col 11: Potential productivity per hour.
        """
        header_map = {
            'potential output1': 'potential_output_m',
            'potential output3': 'potential_output_growth',
            'population (': 'population_16plus_thousands',
            'of which: equilibrium unemployment': 'nairu',
            ' potential employment rate': 'potential_emp_rate',
            'of which: potential participation': 'potential_participation_rate',
            'potential average hours': 'potential_avg_hours',
            'potential productivity per hour': 'potential_productivity',
        }
        records = self._parse_quarterly_sheet(
            'economy', '1.15', header_map,
            header_rows=[4], data_start_row=5,
        )
        return [
            PotentialOutput(
                date=r['date'],
                potential_output_bn=r.get('potential_output_m', 0.0) / 1000.0,
                population_16plus_mn=r.get('population_16plus_thousands', 0.0) / 1000.0,
                nairu=r.get('nairu', 0.0),
                potential_emp_rate=r.get('potential_emp_rate', 0.0),
                potential_avg_hours=r.get('potential_avg_hours', 0.0),
                potential_productivity=r.get('potential_productivity', 0.0),
            )
            for r in records
        ]

    def load_market_assumptions(self) -> List[MarketAssumptions]:
        """Parse Economy Table 1.9: Market-derived assumptions (quarterly)."""
        header_map = {
            'bank rate': 'bank_rate',
            'long-term interest': 'gilt_yield_20y',
            'average mortgage rate': 'average_mortgage_rate',
            'deposit rate': 'deposit_rate',
            'trade-weighted sterling': 'exchange_rate_eri',
            'us$/£ exchange rate': 'usd_exchange_rate',
            'oil prices': 'oil_price_usd',
            'equity prices': 'equity_prices',
        }
        records = self._parse_quarterly_sheet(
            'economy', '1.9', header_map, data_start_row=4,
        )
        return [
            MarketAssumptions(
                date=r['date'],
                bank_rate=r.get('bank_rate', 0.0),
                gilt_yield_20y=r.get('gilt_yield_20y', 0.0),
                average_mortgage_rate=r.get('average_mortgage_rate'),
                exchange_rate_eri=r.get('exchange_rate_eri'),
                oil_price_usd=r.get('oil_price_usd'),
                deposit_rate=r.get('deposit_rate'),
                usd_exchange_rate=r.get('usd_exchange_rate'),
                equity_prices=r.get('equity_prices'),
            )
            for r in records
        ]

    def load_nominal_gdp_components(self) -> List[NominalGDPComponents]:
        """Parse Economy Table 1.2: GDP expenditure components (nominal, quarterly)."""
        header_map = {
            'private consumption': 'private_consumption',
            'government consumption': 'government_consumption',
            'fixed investment': 'fixed_investment',
            'general government': 'general_government_investment',
            'net acquisition of valuables': 'valuables',
            'change in inventories': 'change_in_inventories',
            'exports': 'exports',
            'total final expenditure': 'total_final_expenditure',
            'imports': 'imports',
            'statistical discrepancy': 'statistical_discrepancy',
            'gdp at market prices': 'gdp_market_prices',
        }
        records = self._parse_quarterly_sheet(
            'economy', '1.2', header_map,
            header_rows=[3, 4], data_start_row=5,
        )
        return [
            NominalGDPComponents(
                date=r['date'],
                private_consumption=r.get('private_consumption', 0.0),
                government_consumption=r.get('government_consumption', 0.0),
                fixed_investment=r.get('fixed_investment', 0.0),
                general_government_investment=r.get('general_government_investment', 0.0),
                valuables=r.get('valuables', 0.0),
                change_in_inventories=r.get('change_in_inventories', 0.0),
                exports=r.get('exports', 0.0),
                total_final_expenditure=r.get('total_final_expenditure', 0.0),
                imports=r.get('imports', 0.0),
                statistical_discrepancy=r.get('statistical_discrepancy', 0.0),
                gdp_market_prices=r.get('gdp_market_prices', 0.0),
            )
            for r in records
        ]

    def load_price_indices(self) -> List[PriceIndices]:
        """Parse Economy Table 1.7: Price index levels (quarterly).

        Sheet 1.7 has duplicate headers: growth rates (cols 3-11) and
        levels (cols 12-20). We target the level columns via col_range.
        """
        header_map = {
            'rpi': 'rpi',
            'cpi': 'cpi',
            'cpih': 'cpih',
            'ooh': 'ooh',
            'consumer expenditure deflator': 'pce',
            'gdp deflator': 'pgdp',
        }
        records = self._parse_quarterly_sheet(
            'economy', '1.7', header_map,
            header_rows=[4], data_start_row=5,
            col_range=range(13, 23),
        )
        return [
            PriceIndices(
                date=r['date'],
                rpi=r.get('rpi', 0.0),
                cpi=r.get('cpi', 0.0),
                cpih=r.get('cpih', 0.0),
                ooh=r.get('ooh', 0.0),
                pce=r.get('pce', 0.0),
                pgdp=r.get('pgdp', 0.0),
            )
            for r in records
        ]

    def load_household_income(self) -> List[HouseholdIncome]:
        """Parse Economy Table 1.12: Household disposable income (quarterly, from 2012Q1)."""
        header_map = {
            'labour income': 'labour_income',
            'employee compensation': 'employee_compensation',
            'mixed income': 'mixed_income',
            'employer social contributions': 'employer_social_contributions',
            'non-labour income': 'non_labour_income',
            'net taxes': 'net_benefits_taxes',
            'net benefits': 'net_benefits_taxes',
            'household disposable income': 'household_disposable_income',
        }
        records = self._parse_quarterly_sheet(
            'economy', '1.12', header_map,
            header_rows=[3], data_start_row=4,
        )
        return [
            HouseholdIncome(
                date=r['date'],
                labour_income=r.get('labour_income', 0.0),
                employee_compensation=r.get('employee_compensation', 0.0),
                mixed_income=r.get('mixed_income', 0.0),
                employer_social_contributions=r.get('employer_social_contributions', 0.0),
                non_labour_income=r.get('non_labour_income', 0.0),
                net_benefits_taxes=r.get('net_benefits_taxes', 0.0),
                household_disposable_income=r.get('household_disposable_income', 0.0),
            )
            for r in records
        ]

    def load_balance_of_payments(self) -> List[BalanceOfPayments]:
        """Parse Economy Table 1.8: Balance of payments (quarterly, from 2008Q1)."""
        header_map = {
            'trade balance (': 'trade_balance_pct_gdp',
            'trade balance': 'trade_balance',
            'investment income': 'investment_income_balance',
            'employee income': 'employee_income_balance',
            'transfers balance': 'transfers_balance',
            'current account balance \n(': 'current_account_pct_gdp',
            'current account balance': 'current_account_balance',
        }
        records = self._parse_quarterly_sheet(
            'economy', '1.8', header_map,
            header_rows=[3], data_start_row=4,
        )
        return [
            BalanceOfPayments(
                date=r['date'],
                trade_balance=r.get('trade_balance', 0.0),
                trade_balance_pct_gdp=r.get('trade_balance_pct_gdp', 0.0),
                investment_income_balance=r.get('investment_income_balance', 0.0),
                employee_income_balance=r.get('employee_income_balance', 0.0),
                transfers_balance=r.get('transfers_balance', 0.0),
                current_account_balance=r.get('current_account_balance', 0.0),
                current_account_pct_gdp=r.get('current_account_pct_gdp', 0.0),
            )
            for r in records
        ]

    def load_household_balance_sheet(self) -> List[HouseholdBalanceSheet]:
        """Parse Economy Table 1.11: Household balance sheet (quarterly, from 2012Q1).

        Uses col_range to restrict to household columns (C-H), avoiding
        the PNFC and Lending sections that start at column P.
        """
        header_map = {
            'physical assets': 'physical_assets',
            'financial assets': 'financial_assets',
            'secured liabilities': 'secured_liabilities',
            'other liabilities': 'other_liabilities',
            'liabilities': 'total_liabilities',
            'total net worth': 'total_net_worth',
        }
        records = self._parse_quarterly_sheet(
            'economy', '1.11', header_map,
            header_rows=[4], data_start_row=5,
            col_range=range(3, 9),
        )
        return [
            HouseholdBalanceSheet(
                date=r['date'],
                physical_assets=r.get('physical_assets', 0.0),
                financial_assets=r.get('financial_assets', 0.0),
                total_liabilities=r.get('total_liabilities', 0.0),
                secured_liabilities=r.get('secured_liabilities', 0.0),
                other_liabilities=r.get('other_liabilities', 0.0),
                total_net_worth=r.get('total_net_worth', 0.0),
            )
            for r in records
        ]

    # -- Receipts breakdown (annual) ----------------------------------------

    def load_receipts_breakdown(self) -> List[ReceiptsBreakdown]:
        """
        Parse Receipts summary sheet: Current receipts on a cash basis.

        Sheet is '3.9' (March/November 2025) or '3.8' (March 2026).
        Layout:
            Row 5:  fiscal year headers in columns D–J
            Rows 6–64: tax categories in column B, sub-items in column C
            Row 64: Current receipts (grand total)
        """
        filepath = self.files.get('receipts')
        if not filepath:
            logger.warning("No receipts file found in %s", self.vintage_dir)
            return []

        wb = openpyxl.load_workbook(filepath, data_only=True)

        # Sheet number changed between vintages; validate by title
        ws = None
        for name in ['3.9', '3.8']:
            if name not in wb.sheetnames:
                continue
            candidate = wb[name]
            title = candidate.cell(row=2, column=2).value
            if title and 'current receipts' in str(title).lower():
                ws = candidate
                break
        if ws is None:
            logger.warning("No receipts summary sheet found in %s", filepath)
            wb.close()
            return []

        # Row 5: fiscal year headers starting from column D
        fiscal_years = []
        col = 4  # column D
        while True:
            val = ws.cell(row=5, column=col).value
            if val is None:
                break
            fiscal_years.append(str(val).strip())
            col += 1

        if not fiscal_years:
            logger.warning("No fiscal years found in receipts row 5")
            wb.close()
            return []

        # Search rows for tax category labels in column B
        LABELS = {
            'income tax': 'income_tax_gross',
            'national insurance': 'national_insurance',
            'value added tax': 'vat',
            'corporation tax': 'corporation_tax',
            'fuel duties': 'fuel_duties',
            'business rates': 'business_rates',
            'council tax': 'council_tax',
            'interest and dividends': '_interest_dividends',
            'gross operating surplus': '_gross_operating_surplus',
            'current receipts': 'total_current_receipts',
        }
        sorted_labels = sorted(
            LABELS.items(), key=lambda x: len(x[0]), reverse=True,
        )
        label_rows: Dict[str, List[float]] = {}
        assigned: set = set()

        for row_idx in range(6, 70):
            raw = ws.cell(row=row_idx, column=2).value
            if raw is None:
                continue
            label = str(raw).strip().lower()
            for pattern, field in sorted_labels:
                if field not in assigned and label.startswith(pattern):
                    values = []
                    for i in range(len(fiscal_years)):
                        cell_val = ws.cell(row=row_idx, column=4 + i).value
                        try:
                            values.append(float(cell_val) if cell_val is not None else 0.0)
                        except (ValueError, TypeError):
                            values.append(0.0)
                    label_rows[field] = values
                    assigned.add(field)
                    break

        wb.close()

        # Build ReceiptsBreakdown objects
        results = []
        n = len(fiscal_years)
        zeros = [0.0] * n
        for i, fy in enumerate(fiscal_years):
            income_tax = label_rows.get('income_tax_gross', zeros)[i]
            nics = label_rows.get('national_insurance', zeros)[i]
            vat = label_rows.get('vat', zeros)[i]
            corp_tax = label_rows.get('corporation_tax', zeros)[i]
            fuel = label_rows.get('fuel_duties', zeros)[i]
            brates = label_rows.get('business_rates', zeros)[i]
            ctax = label_rows.get('council_tax', zeros)[i]
            total = label_rows.get('total_current_receipts', zeros)[i]

            int_div = label_rows.get('_interest_dividends', zeros)[i]
            gos = label_rows.get('_gross_operating_surplus', zeros)[i]
            non_tax = int_div + gos

            named_taxes = income_tax + nics + vat + corp_tax + fuel + brates + ctax
            other = total - named_taxes - non_tax

            results.append(ReceiptsBreakdown(
                fiscal_year=fy,
                income_tax_gross=income_tax,
                national_insurance=nics,
                vat=vat,
                corporation_tax=corp_tax,
                fuel_duties=fuel,
                business_rates=brates,
                council_tax=ctax,
                other_taxes=other,
                non_tax_receipts=non_tax,
                total_current_receipts=total,
            ))

        logger.info("Loaded receipts for %d fiscal years", len(results))
        return results

    def load_detailed_receipts(self) -> List[DetailedReceipts]:
        """
        Parse Receipts Table 3.9: all individual tax lines.

        Returns one DetailedReceipts per fiscal year with ~30 fields covering
        every major tax head. Row numbers reference the March 2025 layout.
        """
        filepath = self.files.get('receipts')
        if not filepath:
            logger.warning("No receipts file found in %s", self.vintage_dir)
            return []

        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = None
        for name in ['3.9', '3.8']:
            if name not in wb.sheetnames:
                continue
            candidate = wb[name]
            title = candidate.cell(row=2, column=2).value
            if title and 'current receipts' in str(title).lower():
                ws = candidate
                break
        if ws is None:
            logger.warning("No receipts summary sheet found in %s", filepath)
            wb.close()
            return []

        # Fiscal year headers from row 5
        fiscal_years = []
        col = 4
        while True:
            val = ws.cell(row=5, column=col).value
            if val is None:
                break
            fiscal_years.append(str(val).strip())
            col += 1
        n = len(fiscal_years)
        if n == 0:
            wb.close()
            return []

        def _row_vals(row_idx):
            return [
                float(ws.cell(row=row_idx, column=4 + i).value or 0)
                for i in range(n)
            ]

        # Read specific rows (fixed layout)
        income_tax = _row_vals(7)
        paye = _row_vals(9)
        self_assess = _row_vals(10)
        nics = _row_vals(12)
        vat = _row_vals(13)
        corp_tax = _row_vals(14)
        prt = _row_vals(20)
        fuel = _row_vals(21)
        cgt = _row_vals(22)
        iht = _row_vals(23)
        sdlt = _row_vals(24)
        ated = _row_vals(25)
        stamp_shares = _row_vals(26)
        tobacco = _row_vals(27)
        spirits = _row_vals(28)
        wine = _row_vals(29)
        beer = _row_vals(30)
        apd = _row_vals(31)
        ipt = _row_vals(32)
        ccl = _row_vals(33)
        landfill = _row_vals(34)
        customs = _row_vals(37)
        apprenticeship = _row_vals(42)
        epl = _row_vals(46)
        total_hmrc = _row_vals(51)
        ved = _row_vals(52)
        brates = _row_vals(53)
        ctax = _row_vals(54)
        vat_refunds = _row_vals(55)
        int_div = _row_vals(58)
        gos = _row_vals(62)
        other_rx = _row_vals(63)
        total = _row_vals(64)

        wb.close()

        results = []
        for i, fy in enumerate(fiscal_years):
            results.append(DetailedReceipts(
                fiscal_year=fy,
                income_tax_gross=income_tax[i],
                paye=paye[i],
                self_assessment=self_assess[i],
                national_insurance=nics[i],
                vat=vat[i],
                corporation_tax=corp_tax[i],
                petroleum_revenue_tax=prt[i],
                fuel_duties=fuel[i],
                capital_gains_tax=cgt[i],
                inheritance_tax=iht[i],
                sdlt=sdlt[i],
                ated=ated[i],
                stamp_taxes_shares=stamp_shares[i],
                tobacco_duties=tobacco[i],
                alcohol_duties=spirits[i] + wine[i] + beer[i],
                air_passenger_duty=apd[i],
                insurance_premium_tax=ipt[i],
                climate_change_levy=ccl[i],
                landfill_tax=landfill[i],
                customs_duties=customs[i],
                apprenticeship_levy=apprenticeship[i],
                energy_profits_levy=epl[i],
                total_hmrc=total_hmrc[i],
                vehicle_excise_duties=ved[i],
                business_rates=brates[i],
                council_tax=ctax[i],
                vat_refunds=vat_refunds[i],
                interest_and_dividends=int_div[i],
                gross_operating_surplus=gos[i],
                other_receipts=other_rx[i],
                total_current_receipts=total[i],
            ))

        logger.info("Loaded detailed receipts for %d fiscal years", len(results))
        return results

    # -- Debt interest (annual, from Aggregates 6.16) -----------------------

    def load_debt_interest(self) -> List[DebtInterest]:
        """
        Parse Aggregates Table 6.16: Central government debt interest.

        Layout:
            Row 5:  fiscal year headers in columns C–I
            Row 6:  Debt interest on conventional gilts
            Row 11: Index-linked gilts
            Row 15: NS&I
            Row 16: Other debt interest
            Row 17: Total CG debt interest
        """
        filepath = self.files.get('aggregates')
        if not filepath:
            logger.warning("No aggregates file found in %s", self.vintage_dir)
            return []

        wb = openpyxl.load_workbook(filepath, data_only=True)
        if '6.16' not in wb.sheetnames:
            logger.warning("Sheet '6.16' not found in %s", filepath)
            wb.close()
            return []
        ws = wb['6.16']

        # Row 5: fiscal year headers from column C
        fiscal_years = []
        col = 3
        while True:
            val = ws.cell(row=5, column=col).value
            if val is None:
                break
            fiscal_years.append(str(val).strip())
            col += 1

        if not fiscal_years:
            wb.close()
            return []

        LABELS = {
            'debt interest on conventional': 'conventional_gilts',
            'conventional gilts held in the apf': 'apf_gilts',
            'index-linked': 'index_linked_gilts',
            'debt interest on national savings': 'ns_and_i',
            'other debt interest': 'other',
            'total cg debt interest': 'total',
        }
        sorted_labels = sorted(
            LABELS.items(), key=lambda x: len(x[0]), reverse=True,
        )
        label_rows: Dict[str, List[float]] = {}
        assigned: set = set()

        for row_idx in range(6, 20):
            raw = ws.cell(row=row_idx, column=2).value
            if raw is None:
                continue
            label = str(raw).strip().lower()
            for pattern, field in sorted_labels:
                if field not in assigned and label.startswith(pattern):
                    values = []
                    for i in range(len(fiscal_years)):
                        cell_val = ws.cell(row=row_idx, column=3 + i).value
                        try:
                            values.append(float(cell_val) if cell_val is not None else 0.0)
                        except (ValueError, TypeError):
                            values.append(0.0)
                    label_rows[field] = values
                    assigned.add(field)
                    break

        wb.close()

        results = []
        zeros = [0.0] * len(fiscal_years)
        for i, fy in enumerate(fiscal_years):
            results.append(DebtInterest(
                fiscal_year=fy,
                conventional_gilts=label_rows.get('conventional_gilts', zeros)[i],
                apf_gilts=label_rows.get('apf_gilts', zeros)[i],
                index_linked_gilts=label_rows.get('index_linked_gilts', zeros)[i],
                ns_and_i=label_rows.get('ns_and_i', zeros)[i],
                other=label_rows.get('other', zeros)[i],
                total=label_rows.get('total', zeros)[i],
            ))

        logger.info("Loaded debt interest for %d fiscal years", len(results))
        return results


# ---------------------------------------------------------------------------
# Policy Loader (CSV-based, unchanged)
# ---------------------------------------------------------------------------

class PolicyLoader:
    """Loads policy decisions from CSV files."""

    def __init__(self, policy_database_file: str, new_policy_csv: str = None):
        self.policy_database_file = policy_database_file
        self.new_policy_csv = new_policy_csv

    def load_measures(self) -> List[PolicyMeasure]:
        measures = []
        if not self.new_policy_csv:
            return measures

        try:
            df = pd.read_csv(self.new_policy_csv)
        except FileNotFoundError:
            logger.warning("Policy CSV not found: %s", self.new_policy_csv)
            return measures

        year_cols = [c for c in df.columns if '-' in c and c[0].isdigit()]
        for record in df.to_dict('records'):
            impact = {}
            for col in year_cols:
                try:
                    year_start = int(col.split('-')[0])
                    impact[year_start] = float(record[col])
                except (ValueError, KeyError):
                    continue
            measures.append(PolicyMeasure(
                name=record.get('Measure_Name', 'Unknown'),
                start_year=int(record.get('Start_Year', 0)),
                fiscal_impact_bn=impact,
                uncertainty_rating=record.get('Uncertainty_Rating', 'Medium'),
                labour_supply_impact=float(record.get('Labour_Supply_Impact', 0.0)),
                potential_output_impact=float(record.get('Potential_Output_Impact', 0.0)),
            ))
        return measures
